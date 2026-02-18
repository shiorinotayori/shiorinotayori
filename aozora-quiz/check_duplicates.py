#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
questions-data.xlsx の重複チェックスクリプト

使い方:
    aozora-quiz フォルダ内で実行
    python3 check_duplicates.py

実行すると、E列（本文一節）の重複をチェックして報告します。
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

def check_duplicates(excel_file='questions-data.xlsx'):
    """
    Excelファイルの重複をチェックして、重複セルを赤く塗る
    """
    # ファイルが存在するか確認
    if not os.path.exists(excel_file):
        print(f"❌ エラー: {excel_file} が見つかりません")
        print(f"現在のディレクトリ: {os.getcwd()}")
        print("aozora-quiz フォルダ内で実行してください")
        return
    
    print(f"📖 {excel_file} を読み込み中...")
    
    # Excelファイルを読み込み
    df = pd.read_excel(excel_file, sheet_name='問題データ')
    
    # E列（本文一節）の重複をチェック
    duplicates = df[df.duplicated(subset=['本文一節'], keep=False)]
    
    if len(duplicates) == 0:
        print("\n✅ 重複はありません！")
        return
    
    print(f"\n⚠️  重複が見つかりました: {len(duplicates)}件")
    print("\n重複している問題:")
    
    # 重複をグループ化して表示
    duplicate_groups = df[df.duplicated(subset=['本文一節'], keep=False)].groupby('本文一節')
    
    for text, group in duplicate_groups:
        print(f"\n【重複グループ】")
        for idx, row in group.iterrows():
            print(f"  ID {row['ID']}: {row['作品名']} / {row['著者名']} ({row['難易度']})")
        print(f"  本文: {text[:50]}...")
    
    # Excelファイルに赤色を付ける
    print(f"\n🎨 重複セルを赤く塗ります...")
    
    wb = load_workbook(excel_file)
    ws = wb['問題データ']
    
    # 赤色の塗りつぶし
    red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    
    # 重複している行を特定
    duplicate_indices = df[df.duplicated(subset=['本文一節'], keep=False)].index.tolist()
    
    for idx in duplicate_indices:
        row_num = idx + 2  # Excelは1行目がヘッダー、Pythonは0始まりなので+2
        # E列を赤く塗る
        ws.cell(row=row_num, column=5).fill = red_fill
    
    # 保存
    wb.save(excel_file)
    print(f"✅ {excel_file} を更新しました")
    print(f"   重複している{len(duplicates)}件のセルを赤く塗りました")

if __name__ == '__main__':
    check_duplicates()
